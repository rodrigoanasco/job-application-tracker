from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from job_tracker.models import ApplicationUpdate
from job_tracker.storage import ExcelTracker


class ExcelTrackerTests(unittest.TestCase):
    def test_updates_events_and_current_application_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "tracker.xlsx"
            tracker = ExcelTracker(path)
            updates = [
                _update("msg-1", 0, "received", date(2026, 6, 1)),
                _update("msg-2", 0, "interview", date(2026, 6, 4)),
            ]

            self.assertEqual(tracker.update(updates), 2)
            self.assertEqual(tracker.update(updates), 0)

            workbook = load_workbook(path)
            apps = workbook["Applications"]
            events = workbook["Events"]

            self.assertEqual(events.max_row, 3)
            self.assertEqual(apps.max_row, 2)
            self.assertEqual(apps["A2"].value, "Acme")
            self.assertEqual(apps["C2"].value, "interview")
            self.assertEqual(apps["J2"].value, 2)


def _update(message_id: str, index: int, status: str, event_date: date) -> ApplicationUpdate:
    return ApplicationUpdate(
        company="Acme",
        role="Software Engineer",
        status=status,
        event_date=event_date,
        confidence=0.9,
        notes="test",
        source_message_id=message_id,
        source_subject="Application update",
        source_from="Acme Careers <careers@acme.com>",
        source_date=datetime(2026, 6, event_date.day, tzinfo=timezone.utc),
        event_index=index,
    )


if __name__ == "__main__":
    unittest.main()
