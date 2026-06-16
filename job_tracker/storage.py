from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from .models import ApplicationUpdate

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as error:  # pragma: no cover
    raise RuntimeError("Missing dependency. Run: python -m pip install -r requirements.txt") from error


APPLICATIONS_SHEET = "Applications"
EVENTS_SHEET = "Events"

APPLICATION_HEADERS = [
    "Company",
    "Role",
    "Current Status",
    "Applied Date",
    "Last Update Date",
    "Latest Subject",
    "Latest From",
    "Confidence",
    "Notes",
    "Event Count",
    "Key",
]

EVENT_HEADERS = [
    "Event ID",
    "Message ID",
    "Email Date",
    "Company",
    "Role",
    "Status",
    "Event Date",
    "Subject",
    "From",
    "Confidence",
    "Notes",
    "Key",
]


class ExcelTracker:
    def __init__(self, path: Path) -> None:
        self.path = path

    def update(self, updates: list[ApplicationUpdate]) -> int:
        if not updates:
            self._load_or_create().save(self.path) if self.path.exists() else None
            return 0

        workbook = self._load_or_create()
        events = workbook[EVENTS_SHEET]
        existing_event_ids = self._existing_event_ids(events)

        added = 0
        for update in updates:
            if update.event_id in existing_event_ids:
                continue
            events.append(_event_row(update))
            existing_event_ids.add(update.event_id)
            added += 1

        if added:
            self._rebuild_applications(workbook)
            self._format(workbook[APPLICATIONS_SHEET])
            self._format(workbook[EVENTS_SHEET])
            self.path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(self.path)

        return added

    def ensure_exists(self) -> None:
        workbook = self._load_or_create()
        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self.path)

    def _load_or_create(self):
        if self.path.exists():
            workbook = load_workbook(self.path)
        else:
            workbook = Workbook()

        applications = _ensure_sheet(workbook, APPLICATIONS_SHEET, APPLICATION_HEADERS)
        events = _ensure_sheet(workbook, EVENTS_SHEET, EVENT_HEADERS)
        self._format(applications)
        self._format(events)
        return workbook

    def _existing_event_ids(self, events: Worksheet) -> set[str]:
        ids: set[str] = set()
        for row in events.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                ids.add(str(row[0]))
        return ids

    def _rebuild_applications(self, workbook) -> None:
        applications = workbook[APPLICATIONS_SHEET]
        events = workbook[EVENTS_SHEET]

        if applications.max_row > 1:
            applications.delete_rows(2, applications.max_row - 1)

        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in events.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            event = dict(zip(EVENT_HEADERS, row))
            grouped[str(event["Key"])].append(event)

        for key in sorted(grouped):
            group = grouped[key]
            group.sort(key=lambda item: _date_sort_key(item.get("Event Date"), item.get("Email Date")))
            latest = group[-1]
            applied_date = _first_applied_date(group)
            applications.append(
                [
                    latest.get("Company"),
                    latest.get("Role"),
                    latest.get("Status"),
                    applied_date,
                    _display_date(latest.get("Event Date") or latest.get("Email Date")),
                    latest.get("Subject"),
                    latest.get("From"),
                    latest.get("Confidence"),
                    latest.get("Notes"),
                    len(group),
                    key,
                ]
            )

    def _format(self, sheet: Worksheet) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        widths = {
            "A": 24,
            "B": 30,
            "C": 18,
            "D": 16,
            "E": 18,
            "F": 46,
            "G": 32,
            "H": 12,
            "I": 48,
            "J": 12,
            "K": 40,
            "L": 40,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width


def _ensure_sheet(workbook, name: str, headers: list[str]) -> Worksheet:
    if name in workbook.sheetnames:
        sheet = workbook[name]
    else:
        sheet = workbook.active if workbook.active.max_row == 1 and workbook.active.max_column == 1 and workbook.active["A1"].value is None else workbook.create_sheet(name)
        sheet.title = name

    if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
        for index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=index, value=header)
        return sheet

    existing_headers = [cell.value for cell in sheet[1]]
    if existing_headers[: len(headers)] != headers:
        if sheet.max_row > 1 or any(cell.value for cell in sheet[1]):
            raise ValueError(f"Sheet {name} has unexpected columns. Move it aside or update it manually.")
        for index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=index, value=header)
    return sheet


def _event_row(update: ApplicationUpdate) -> list[Any]:
    event_date = update.event_date or update.source_date.date()
    return [
        update.event_id,
        update.source_message_id,
        update.source_date.replace(tzinfo=None),
        update.company,
        update.role,
        update.status,
        event_date,
        update.source_subject,
        update.source_from,
        update.confidence,
        update.notes,
        update.application_key,
    ]


def _first_applied_date(group: list[dict[str, Any]]) -> Any:
    candidates = [
        item.get("Event Date")
        for item in group
        if item.get("Status") in {"applied", "received"} and item.get("Event Date")
    ]
    if not candidates:
        return None
    return _display_date(min(candidates, key=lambda value: _date_sort_key(value, value)))


def _display_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date()
    return value


def _date_sort_key(primary: Any, fallback: Any) -> datetime:
    value = primary or fallback
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            pass
    return datetime.min
